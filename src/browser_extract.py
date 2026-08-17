"""
Browser Extraction & Attachment Download Module (src.browser_extract)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

Selenium-based module for extracting attachments from FTIR detail pages
behind the SIFT Maruti internal quality portal.

Strategy
--------
*  Uses a **persistent Chrome user-data-dir profile** so the operator
   only needs to log into SIFT Maruti **once** in their default browser;
   subsequent automated runs reuse the saved session cookies & tokens.
*  The script **never** handles login credentials itself.
*  Downloads are performed via **requests** using the browser session's
   cookies, which is faster and more reliable than Selenium's built-in
   download manager for binary files.

Usage::

    # Test against a single FTIR page:
    python -m src.browser_extract "https://sift.marutisuzuki.com/ftir/12345"

    # With a custom profile directory:
    python -m src.browser_extract "https://sift.marutisuzuki.com/ftir/12345" \\
        --profile /path/to/chrome-profile
"""

import os
import sys
import time
import logging
import argparse
import mimetypes
import re
from typing import Dict, List, Optional, Any, Union
from urllib.parse import urlparse, urljoin, unquote

import requests
from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    WebDriverException,
)

logger = logging.getLogger(__name__)

# Default persistent profile location (platform-aware)
_DEFAULT_PROFILE_DIR = os.path.join(
    os.path.expanduser("~"), ".ftir_sbpr_tool", "browser_profile"
)

# ---------------------------------------------------------------------------
#  CONFIGURATION
# ---------------------------------------------------------------------------
# Portal URL is read from config/portal_url.txt (gitignored so git pull
# never overwrites your URL).  Falls back to this hardcoded default.
DEFAULT_SIFT_PORTAL_URL = "https://sift.bizapps.suzuki/sift/"


def _read_config_file_clean(filepath: str) -> str:
    """Read a config file trying multiple encodings and cleaning whitespace/comments/BOM."""
    if not os.path.isfile(filepath):
        return ""
    for enc in ["utf-8-sig", "utf-16", "utf-8", "latin-1", "cp1252"]:
        try:
            with open(filepath, "r", encoding=enc) as f:
                content = f.read()
            cleaned = content.replace("\x00", "").replace("\ufeff", "").strip()
            for line in cleaned.splitlines():
                line = line.strip().strip("'\"")
                if line and not line.startswith("#"):
                    return line
        except Exception:
            continue
    return ""


def _load_portal_url() -> str:
    """Read portal URL from config/portal_url.txt, falling back to default SIFT URL."""
    _project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = [
        os.path.join(_project_root, "config", "portal_url.txt"),
        os.path.join(os.getcwd(), "config", "portal_url.txt"),
        "config/portal_url.txt",
    ]
    for p in candidates:
        url = _read_config_file_clean(p)
        if url and url.startswith("http") and "YOUR_PORTAL_URL" not in url and "INSERT_URL" not in url:
            return url
    return DEFAULT_SIFT_PORTAL_URL


def _load_attachment_template() -> str:
    """Read exact attachment URL template from config/attachment_url_template.txt."""
    _project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = [
        os.path.join(_project_root, "config", "attachment_url_template.txt"),
        os.path.join(os.getcwd(), "config", "attachment_url_template.txt"),
        "config/attachment_url_template.txt",
    ]
    for config_path in candidates:
        if os.path.isfile(config_path):
            for enc in ["utf-8-sig", "utf-16", "utf-8", "latin-1", "cp1252"]:
                try:
                    with open(config_path, "r", encoding=enc) as f:
                        for line in f:
                            line = line.replace("\x00", "").replace("\ufeff", "").strip()
                            if line and not line.startswith("#"):
                                if "http" in line and "{FTIR_NO}" in line:
                                    return line
                except Exception:
                    continue
    return ""


PORTAL_SEARCH_URL = _load_portal_url()
logger.info(f"✓ Portal URL loaded: {PORTAL_SEARCH_URL}")

ATTACHMENT_URL_TEMPLATE = _load_attachment_template()
if ATTACHMENT_URL_TEMPLATE:
    logger.info("✓ Exact Attachment URL Template loaded from config!")

# Delay (seconds) between consecutive network requests to avoid
# hammering the internal server.
_REQUEST_DELAY = 1.5

# Page-load timeout for Selenium waits (seconds).
# Set extremely high (5 minutes) so that if the user hits a login screen
# and has to type passwords/OTP, the script doesn't time out and crash
# while they are still typing!
_PAGE_TIMEOUT = 300

# Common attachment file extensions we care about.
_ATTACHMENT_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif",
    ".webp", ".heic",
    ".pdf",
    ".mp4", ".avi", ".mov", ".wmv", ".mkv", ".3gp", ".webm",
    ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".zip", ".rar", ".7z",
}


# ---------------------------------------------------------------------------
#  Driver setup
# ---------------------------------------------------------------------------

def get_driver(
    profile_dir: Optional[str] = None,
    headless: bool = False,
) -> webdriver.Edge:
    """
    Initialise the Selenium WebDriver.
    Strictly launches Microsoft Edge for enterprise compatibility.

    Parameters
    ----------
    profile_dir : str, optional
        Filesystem path to the Edge user-data directory.
        Defaults to ``~/.ftir_sbpr_tool/browser_profile``.
    headless : bool
        If True, launch Edge in headless mode (no visible window).
        Defaults to False so the user can see the browser.

    Returns
    -------
    webdriver.Edge
    """
    if profile_dir is None:
        profile_dir = _DEFAULT_PROFILE_DIR + "_edge"

    os.makedirs(profile_dir, exist_ok=True)

    edge_options = EdgeOptions()

    # Persistent profile — preserves login sessions across runs
    edge_options.add_argument(f"--user-data-dir={os.path.abspath(profile_dir)}")

    if headless:
        # True headless breaks enterprise portals. Move window off-screen instead to hide it.
        edge_options.add_argument("--window-position=-32000,-32000")

    # Suppress noisy DevTools logging
    edge_options.add_argument("--log-level=3")
    edge_options.add_experimental_option("excludeSwitches", ["enable-logging", "enable-automation"])
    edge_options.add_experimental_option("useAutomationExtension", False)

    # Reasonable window size for element visibility
    edge_options.add_argument("--window-size=1920,1080")

    # Disable pop-up blocker so download dialogs don't interfere
    edge_options.add_argument("--disable-popup-blocking")

    # Set default download directory and disable prompt dialogs
    _project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    default_download_dir = os.path.join(_project_root, "temp_downloads")
    os.makedirs(default_download_dir, exist_ok=True)
    prefs = {
        "download.default_directory": os.path.abspath(default_download_dir),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "plugins.always_open_pdf_externally": True,
    }
    # Enable Performance Logging if supported
    try:
        edge_options.set_capability("goog:loggingPrefs", {"performance": "ALL"})
        edge_options.set_capability("ms:loggingPrefs", {"performance": "ALL"})
    except Exception:
        pass

    try:
        driver = webdriver.Edge(options=edge_options)
        logger.info(f"Edge driver started with profile: {profile_dir}")
    except WebDriverException as e:
        raise RuntimeError(
            f"Could not start Microsoft Edge browser.\n"
            f"Make sure Edge is installed and up to date.\n"
            f"Error: {e}"
        )

    driver.set_page_load_timeout(_PAGE_TIMEOUT)
    driver.implicitly_wait(5)
    return driver


# ---------------------------------------------------------------------------
#  Page extraction
# ---------------------------------------------------------------------------

def _get_browser_cookies_for_requests(driver: webdriver.Edge) -> dict:
    """Convert Selenium cookies into a dict usable by ``requests``."""
    return {c["name"]: c["value"] for c in driver.get_cookies()}


def _looks_like_attachment_url(href: str) -> bool:
    """Heuristic: does this URL look like a downloadable attachment?"""
    if not href:
        return False
    parsed = urlparse(href)
    path_lower = parsed.path.lower()
    # Direct file extension match
    if any(path_lower.endswith(ext) for ext in _ATTACHMENT_EXTENSIONS):
        return True
    # Common download-servlet patterns
    download_patterns = [
        "download", "attach", "getfile", "blob", "media",
        "upload", "file", "resource", "content",
    ]
    if any(kw in path_lower for kw in download_patterns):
        return True
    # Query-param based downloads (e.g. ?action=download&id=...)
    query_lower = parsed.query.lower()
    if any(kw in query_lower for kw in ["download", "attach", "file"]):
        return True
    return False


def _extract_sift_page_entities(driver: webdriver.Edge) -> List[Dict[str, Any]]:
    """
    Extract exact attachment records from SIFT's syqaa090FindWebAttachmentEntity hidden fields
    and comOnFtirAttachmentFileLinkClicked / ftirWebThumbnail elements across all frames.
    """
    js_extract_entities = """
    var results = [];
    var seenKeys = {};

    function scanDocument(doc, win) {
        if (!doc) return;
        
        // 1. Scan hidden inputs for syqaa090FindWebAttachmentEntity
        var inputs = doc.querySelectorAll("input[name*='syqaa090FindWebAttachmentEntity']");
        var entities = {};
        for (var i = 0; i < inputs.length; i++) {
            var name = inputs[i].name;
            var val = inputs[i].value;
            var m = name.match(/syqaa090FindWebAttachmentEntity\s*\[\s*(\d+)\s*\]\.(\w+)/);
            if (m) {
                var idx = m[1];
                var field = m[2];
                if (!entities[idx]) entities[idx] = {};
                entities[idx][field] = val;
            }
        }
        for (var k in entities) {
            var ent = entities[k];
            var docId = ent.f1Id || "";
            var seq = ent.f1Seq || "";
            var cat = ent.f1FileCategory || "1";
            var fname = ent.f1Name || "";
            var fsize = ent.f1Size || "";
            var ftype = ent.f1Type || "";
            if (docId && seq) {
                var uniqueKey = docId + "_" + cat + "_" + seq;
                if (!seenKeys[uniqueKey]) {
                    seenKeys[uniqueKey] = true;
                    results.push({
                        "doc_id": docId,
                        "file_category": cat,
                        "file_sequence": seq,
                        "filename": fname,
                        "file_size": fsize,
                        "file_type": ftype,
                        "source": "hidden_entity"
                    });
                }
            }
        }

        // 2. Scan a tags with onclick containing comOnFtirAttachmentFileLinkClicked
        var aTags = doc.querySelectorAll("a[onclick*='comOnFtirAttachmentFileLinkClicked']");
        for (var a = 0; a < aTags.length; a++) {
            var oc = aTags[a].getAttribute("onclick") || "";
            var am = oc.match(/comOnFtirAttachmentFileLinkClicked\s*\(\s*['"]([^'"]+)['"]\s*,\s*['"]([^'"]+)['"]\s*,\s*['"]([^'"]+)['"]/);
            if (am) {
                var docId = am[1];
                var cat = am[2];
                var seq = am[3];
                var fname = aTags[a].innerText.trim();
                var uniqueKey = docId + "_" + cat + "_" + seq;
                if (!seenKeys[uniqueKey]) {
                    seenKeys[uniqueKey] = true;
                    results.push({
                        "doc_id": docId,
                        "file_category": cat,
                        "file_sequence": seq,
                        "filename": fname,
                        "file_size": "",
                        "file_type": "",
                        "source": "onclick_handler"
                    });
                }
            }
        }

        // 3. Scan img tags with ftirWebThumbnail or ftirWeb
        var imgTags = doc.querySelectorAll("img[src*='ftirWeb']");
        for (var im = 0; im < imgTags.length; im++) {
            var src = imgTags[im].src || "";
            var docIdM = src.match(/[?&]documentId=([^&]+)/i);
            var seqM = src.match(/[?&]fileSequence=([^&]+)/i);
            var catM = src.match(/[?&]fileCategory=([^&]+)/i);
            if (docIdM && seqM) {
                var docId = docIdM[1];
                var seq = seqM[1];
                var cat = catM ? catM[1] : "1";
                var uniqueKey = docId + "_" + cat + "_" + seq;
                if (!seenKeys[uniqueKey]) {
                    seenKeys[uniqueKey] = true;
                    results.push({
                        "doc_id": docId,
                        "file_category": cat,
                        "file_sequence": seq,
                        "filename": "",
                        "file_size": "",
                        "file_type": "",
                        "source": "img_src"
                    });
                }
            }
        }
    }

    try {
        scanDocument(document, window);
    } catch(e) {}
    return results;
    """
    all_entities = []
    seen = set()

    def _collect_from_context():
        try:
            res = driver.execute_script(js_extract_entities)
            if res and isinstance(res, list):
                for item in res:
                    key = (item.get("doc_id"), item.get("file_category"), item.get("file_sequence"))
                    if key not in seen:
                        seen.add(key)
                        all_entities.append(item)
        except Exception:
            pass

    # Collect from main context
    driver.switch_to.default_content()
    _collect_from_context()

    # Collect from all frames / iframes
    try:
        frames = driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame")
        for idx in range(len(frames)):
            try:
                curr_frames = driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame")
                if idx >= len(curr_frames):
                    continue
                driver.switch_to.frame(curr_frames[idx])
                _collect_from_context()
                driver.switch_to.parent_frame()
            except Exception:
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
    except Exception:
        pass
    finally:
        driver.switch_to.default_content()

    return all_entities


def _extract_sift_page_metadata(driver: webdriver.Edge) -> Dict[str, Any]:
    """
    Extract structured metadata directly from the live SIFT detail page DOM.
    """
    js_extract_meta = """
    function getVal(name) {
        var el = document.querySelector("[name='" + name + "']");
        if (el) return (el.value || el.innerText || "").trim();
        return "";
    }
    return {
        "subject": getVal("syqaa090FindTFtirViewEntity.g1SubjectEn") || getVal("syqaa090FindTFtirViewEntity.f1FaultSubject"),
        "customer_complaint": getVal("syqaa090FindTFtirViewEntity.f1FaultProposal"),
        "incident_condition": getVal("syqaa090FindTFtirViewEntity.f1FaultBecame"),
        "checked_contents": getVal("syqaa090FindTFtirViewEntity.f1FaultCheck"),
        "checked_results": getVal("syqaa090FindTFtirViewEntity.f1FaultResult"),
        "casual_parts_number": getVal("syqaa090FindTFtirViewEntity.g1CausalPartsNo"),
        "casual_parts_name": getVal("syqaa090FindTFtirViewEntity.g1CausalPartsNameEn"),
        "product_model_code": getVal("syqaa090FindTFtirViewEntity.g1ProductModelCode"),
        "sales_model_code": getVal("syqaa090FindTFtirViewEntity.g1SalesModelCode"),
        "vin": getVal("syqaa090FindTFtirViewEntity.f1Vin"),
        "mileage": getVal("syqaa090FindTFtirViewEntity.f1MileageTimeView"),
        "diagnosis_code": getVal("syqaa090FindTFtirViewEntity.f1FaultDtc"),
        "date_reported": getVal("syqaa090FindTFtirViewEntity.f1ReportDate"),
        "date_of_incident": getVal("syqaa090FindTFtirViewEntity.f1FailureDate")
    };
    """
    metadata = {}
    try:
        driver.switch_to.default_content()
        meta = driver.execute_script(js_extract_meta)
        if meta and isinstance(meta, dict):
            metadata = {k: v for k, v in meta.items() if v}
    except Exception as e:
        logger.debug(f"DOM metadata extraction error: {e}")
    finally:
        driver.switch_to.default_content()
    return metadata


def _download_via_browser_fetch(driver: webdriver.Edge, url: str) -> Optional[bytes]:
    """
    Download binary resource directly inside the active browser session via fetch()
    and FileReader, perfectly preserving all session cookies, SSO auth, and TLS state.
    """
    import base64
    js = """
    var uri = arguments[0];
    var callback = arguments[1];
    fetch(uri, { credentials: 'include' })
        .then(function(r) {
            if (!r.ok) throw new Error('HTTP ' + r.status);
            return r.blob();
        })
        .then(function(blob) {
            var reader = new FileReader();
            reader.onloadend = function() {
                callback(reader.result);
            };
            reader.readAsDataURL(blob);
        })
        .catch(function(err) {
            callback(null);
        });
    """
    try:
        data_url = driver.execute_async_script(js, url)
        if data_url and isinstance(data_url, str) and "," in data_url:
            b64_content = data_url.split(",", 1)[1]
            raw_bytes = base64.b64decode(b64_content)
            if len(raw_bytes) > 100:
                return raw_bytes
    except Exception as e:
        logger.debug(f"Browser in-session fetch failed for {url}: {e}")
    return None


def extract_ftir_page(driver: webdriver.Edge, url: str, ftir_no: str = None) -> Dict[str, Any]:
    """
    Navigate to an FTIR detail page and extract subject text, rich metadata,
    and all attachment records directly from the live DOM entities.
    """
    if url and url.startswith("http"):
        try:
            cur_url = driver.current_url.lower()
            if "syqaa090" not in cur_url and "syqaa710" in cur_url:
                logger.info(f"Navigating to FTIR page: {url}")
                driver.get(url)
        except Exception as e:
            logger.warning(f"Navigation issue: {e}")

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
    except TimeoutException:
        pass

    _wait_for_login_if_needed(driver)

    page_title = driver.title or ""
    base_url = driver.current_url

    # 1. Extract rich metadata from live page DOM
    dom_metadata = _extract_sift_page_metadata(driver)
    subject_text = dom_metadata.get("subject") or None

    # Fallback subject extraction if DOM query returned none
    if not subject_text:
        subject_selectors = [
            "//td[contains(translate(text(),'SUBJECT','subject'),'subject')]/following-sibling::td",
            "//th[contains(translate(text(),'SUBJECT','subject'),'subject')]/following-sibling::td",
            "//label[contains(translate(text(),'SUBJECT','subject'),'subject')]/following-sibling::*",
            "//span[contains(translate(text(),'SUBJECT','subject'),'subject')]/parent::*/following-sibling::*",
            "//textarea[contains(@name,'Subject')]",
            "//textarea[contains(@name,'subject')]",
        ]
        for xpath in subject_selectors:
            try:
                el = driver.find_element(By.XPATH, xpath)
                txt = (el.text or el.get_attribute("value") or "").strip()
                if txt and len(txt) > 3:
                    subject_text = txt
                    break
            except NoSuchElementException:
                continue

    # 2. Extract exact attachment entities from SIFT hidden inputs and onclick links
    logger.info("Scanning live SIFT DOM for Syqaa090 WebAttachment entities...")
    entities = _extract_sift_page_entities(driver)

    attachment_items: List[Dict[str, Any]] = []
    attachment_urls: List[str] = []
    seen_urls: set = set()

    # Determine base SIFT URL for building file/thumbnail links
    base_sift = "https://sift.bizapps.suzuki/sift"
    if "/sift" in base_url.lower():
        idx = base_url.lower().find("/sift")
        base_sift = base_url[:idx] + "/sift"

    for ent in entities:
        doc_id = ent["doc_id"]
        seq = ent["file_sequence"]
        cat = ent["file_category"]
        fname = ent["filename"]
        
        full_url = f"{base_sift}/ftirWebFile.do?documentId={doc_id}&fileSequence={seq}&fileCategory={cat}&timeZoneOffset=-330"
        thumb_url = f"{base_sift}/ftirWebThumbnail.do?documentId={doc_id}&fileSequence={seq}&fileCategory={cat}&timeZoneOffset=-330"

        if not fname:
            fname = f"{doc_id}_{cat}_{seq}.jpg"

        if full_url not in seen_urls:
            seen_urls.add(full_url)
            attachment_urls.append(full_url)
            attachment_items.append({
                "url": full_url,
                "fallback_url": thumb_url,
                "filename": fname,
                "doc_id": doc_id,
                "file_sequence": seq,
                "file_category": cat,
                "source": ent.get("source", "sift_entity"),
            })
            logger.info(f"  ✓ Found FTIR attachment entity: {fname} (docId={doc_id}, seq={seq})")

    # 3. Fallback scan for generic <a> / <img> tags if no entities were detected
    if not attachment_items:
        logger.info("Scanning for generic image links/thumbnails...")
        for a_tag in driver.find_elements(By.TAG_NAME, "a"):
            href = a_tag.get_attribute("href")
            if href and href not in seen_urls:
                href_lower = href.lower()
                if "ftirweb" in href_lower or "webfile" in href_lower:
                    seen_urls.add(href)
                    attachment_urls.append(href)
                    attachment_items.append({
                        "url": href,
                        "fallback_url": re.sub(r'ftirWebFile\.do', 'ftirWebThumbnail.do', href, flags=re.IGNORECASE),
                        "filename": (a_tag.text or "").strip() or None,
                        "source": "generic_link"
                    })

        for img_tag in driver.find_elements(By.TAG_NAME, "img"):
            src = img_tag.get_attribute("src")
            if src and src not in seen_urls and not src.startswith("data:"):
                if "ftirwebthumbnail.do" in src.lower():
                    full_src = re.sub(r'ftirWebThumbnail\.do', 'ftirWebFile.do', src, flags=re.IGNORECASE)
                    if full_src not in seen_urls:
                        seen_urls.add(full_src)
                        attachment_urls.append(full_src)
                        attachment_items.append({
                            "url": full_src,
                            "fallback_url": src,
                            "filename": None,
                            "source": "thumbnail_img"
                        })

    # 4. Fallback pattern guesser if still no images found
    if not attachment_items and ftir_no:
        logger.info(f"Generating predictable SIFT attachment URLs for {ftir_no}...")
        for category in range(1, 4):
            for sequence in range(1, 6):
                full_url = f"{base_sift}/ftirWebFile.do?documentId={ftir_no}&fileCategory={category}&fileSequence={sequence}&timeZoneOffset=-330"
                thumb_url = f"{base_sift}/ftirWebThumbnail.do?documentId={ftir_no}&fileCategory={category}&fileSequence={sequence}&timeZoneOffset=-330"
                if full_url not in seen_urls:
                    seen_urls.add(full_url)
                    attachment_urls.append(full_url)
                    attachment_items.append({
                        "url": full_url,
                        "fallback_url": thumb_url,
                        "filename": f"{ftir_no}_cat{category}_seq{sequence}.jpg",
                        "source": "predicted_pattern"
                    })

    # Resolve relative URLs to absolute
    attachment_urls = [urljoin(base_url, u) for u in attachment_urls]

    logger.info(
        f"Extracted {len(attachment_items)} attachment item(s) from FTIR page "
        f"(subject: {'found' if subject_text else 'not found'})"
    )

    return {
        "page_url": base_url,
        "subject_text": subject_text,
        "attachment_urls": attachment_urls,
        "attachment_items": attachment_items,
        "page_title": page_title,
        "ftir_metadata": dom_metadata,
    }


# ---------------------------------------------------------------------------
#  Window management helpers (ported from proven reference script)
# ---------------------------------------------------------------------------

def _go_to_main_window(driver: webdriver.Edge) -> None:
    """Switch back to the first (main) browser window and default content."""
    try:
        driver.switch_to.window(driver.window_handles[0])
        driver.switch_to.default_content()
    except Exception as e:
        logger.warning(f"Could not switch to main window: {e}")


def _close_extra_windows(driver: webdriver.Edge) -> None:
    """Close all browser windows except the first (main) one."""
    try:
        main_window = driver.window_handles[0]
        for handle in driver.window_handles[1:]:
            try:
                if handle in driver.window_handles:
                    driver.switch_to.window(handle)
                    driver.close()
                    time.sleep(0.5)
            except Exception as e:
                logger.debug(f"Window close issue ignored: {e}")
        if main_window in driver.window_handles:
            driver.switch_to.window(main_window)
            driver.switch_to.default_content()
    except Exception as e:
        logger.warning(f"Cleanup issue ignored: {e}")


def _wait_for_login_if_needed(driver: webdriver.Edge) -> None:
    """Detect SSO login screen and pause for manual login if present."""
    current_url = driver.current_url.lower()
    has_login_url = "login" in current_url or "sso" in current_url or "auth" in current_url
    has_sso_fields = (
        len(driver.find_elements(By.ID, "username")) > 0
        or len(driver.find_elements(By.ID, "password")) > 0
        or len(driver.find_elements(By.ID, "cal-login-button")) > 0
        or len(driver.find_elements(By.CSS_SELECTOR, "input[type='password']")) > 0
    )
    if has_login_url or has_sso_fields:
        logger.info(f"⚠️ SSO Login screen detected. Please log in manually. Waiting up to {_PAGE_TIMEOUT}s...")
        try:
            WebDriverWait(driver, _PAGE_TIMEOUT).until(
                lambda d: (
                    "login" not in d.current_url.lower()
                    and len(d.find_elements(By.ID, "cal-login-button")) == 0
                    and len(d.find_elements(By.CSS_SELECTOR, "input[type='password']")) == 0
                )
            )
            logger.info("✓ Login appears successful!")
            time.sleep(2)
        except TimeoutException:
            logger.error("Login timed out. Proceeding anyway.")


# ---------------------------------------------------------------------------
#  Quick Search → FTIR Detail Page via Popup Windows
# ---------------------------------------------------------------------------

def _navigate_to_ftir_detail_via_quick_search(
    driver: webdriver.Edge,
    ftir_no: str,
    portal_url: Optional[str] = None,
) -> bool:
    """
    Navigate to the FTIR detail page using the exact popup-window chain
    proven to work on the live SIFT portal.

    Flow:
      1. Ensure portal is loaded (Menu frameset SYQAA710E01SFnd.do)
      2. Dynamically scan frames for the Menu tree and 'QUICK SEARCH'
      3. Click Quick Search → new popup window opens
      4. Enter FTIR number in #txtSel0, click #searchbtn → FTIR detail popup opens
      5. Switch to and confirm FTIR detail window

    Returns True if we successfully land on the FTIR detail window.
    """
    if not portal_url or not portal_url.startswith("http") or "YOUR_PORTAL_URL" in portal_url or "INSERT_URL" in portal_url:
        portal_url = PORTAL_SEARCH_URL or DEFAULT_SIFT_PORTAL_URL

    logger.info(f"FTIR {ftir_no}: Navigating via Quick Search popup chain on portal {portal_url}...")

    # ── Step 1: Navigate to portal (or reuse existing portal session) ───
    _go_to_main_window(driver)
    current_url = ""
    try:
        current_url = driver.current_url.lower()
    except Exception:
        pass

    if "sift" not in current_url and "bizapps.suzuki" not in current_url:
        logger.info(f"FTIR {ftir_no}: Loading portal base URL: {portal_url}")
        driver.get(portal_url)
        time.sleep(2)

    # ── Check for SSO login ─────────────────────────────────────────────
    _wait_for_login_if_needed(driver)

    # ── Step 2: Switch to Menu frame (dynamically locate frame with Quick Search) ──
    _go_to_main_window(driver)
    driver.switch_to.default_content()
    
    found_menu_frame = False
    try:
        frames = driver.find_elements(By.TAG_NAME, "frame") + driver.find_elements(By.TAG_NAME, "iframe")
        for idx, frame in enumerate(frames):
            try:
                driver.switch_to.default_content()
                driver.switch_to.frame(frame)
                # Check for Quick Search keyword or element
                qs_matches = driver.find_elements(By.XPATH, "//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'quick search')]")
                if qs_matches:
                    logger.info(f"FTIR {ftir_no}: Located 'QUICK SEARCH' in frame index {idx}")
                    found_menu_frame = True
                    break
            except Exception:
                continue
    except Exception as e:
        logger.debug(f"Frame scanning exception: {e}")

    if not found_menu_frame:
        # Fallback to frame index 1 (menuFrame)
        driver.switch_to.default_content()
        try:
            driver.switch_to.frame(1)
            found_menu_frame = True
            logger.info(f"FTIR {ftir_no}: Switched to default menuFrame (index 1)")
        except Exception:
            try:
                driver.switch_to.frame("menuFrame")
                found_menu_frame = True
                logger.info(f"FTIR {ftir_no}: Switched to frame by name 'menuFrame'")
            except Exception as e:
                logger.error(f"FTIR {ftir_no}: Could not switch to menuFrame: {e}")
                return False

    # ── Step 3: Click Quick Search ──────────────────────────────────────
    old_windows = list(driver.window_handles)
    quick_search_clicked = False
    
    qs_selectors = [
        "//div[@id='group2content']//table[@class='NoBorderTable']//tr/td[2][contains(., 'QUICK SEARCH')]",
        "//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'quick search')]",
        "//*[contains(translate(@title, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'quick search')]",
        "//a[contains(., 'QUICK SEARCH') or contains(., 'Quick Search')]",
        "//td[contains(., 'QUICK SEARCH') or contains(., 'Quick Search')]",
    ]

    for xpath in qs_selectors:
        try:
            elements = driver.find_elements(By.XPATH, xpath)
            for el in elements:
                if el.is_displayed():
                    try:
                        driver.execute_script("arguments[0].click();", el)
                        logger.info(f"FTIR {ftir_no}: Clicked Quick Search via JS")
                        quick_search_clicked = True
                        break
                    except Exception:
                        el.click()
                        logger.info(f"FTIR {ftir_no}: Clicked Quick Search via native click")
                        quick_search_clicked = True
                        break
            if quick_search_clicked:
                break
        except Exception:
            continue

    if not quick_search_clicked:
        logger.error(f"FTIR {ftir_no}: Could not find or click 'QUICK SEARCH' element in menu")
        driver.switch_to.default_content()
        return False

    driver.switch_to.default_content()

    # ── Step 4: Switch to Quick Search popup window ─────────────────────
    quick_search_window = None
    start_wait = time.time()
    while time.time() - start_wait < 20:
        current_handles = driver.window_handles
        new_handles = [h for h in current_handles if h not in old_windows]
        if new_handles:
            quick_search_window = new_handles[0]
            break
        # Also check if any existing window has the #txtSel0 element
        for h in current_handles:
            if len(driver.window_handles) > 0 and h != driver.window_handles[0]:
                try:
                    driver.switch_to.window(h)
                    if len(driver.find_elements(By.ID, "txtSel0")) > 0:
                        quick_search_window = h
                        break
                except Exception:
                    continue
        if quick_search_window:
            break
        time.sleep(1)

    if not quick_search_window:
        logger.error(f"FTIR {ftir_no}: Quick Search popup window did not appear")
        return False

    driver.switch_to.window(quick_search_window)
    logger.info(f"FTIR {ftir_no}: Switched to Quick Search popup window ({quick_search_window})")
    time.sleep(1)

    # ── Step 5: Enter FTIR number in #txtSel0 ──────────────────────────
    try:
        search_box = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "txtSel0"))
        )
        search_box.clear()
        search_box.send_keys(str(ftir_no))
        logger.info(f"FTIR {ftir_no}: Entered '{ftir_no}' into #txtSel0")
    except Exception as e:
        logger.error(f"FTIR {ftir_no}: Could not locate search box #txtSel0: {e}")
        return False

    # ── Step 6: Click #searchbtn ────────────────────────────────────────
    pre_search_windows = list(driver.window_handles)
    search_clicked = False
    try:
        search_btn = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "searchbtn"))
        )
        try:
            driver.execute_script("arguments[0].click();", search_btn)
            search_clicked = True
        except Exception:
            search_btn.click()
            search_clicked = True
        logger.info(f"FTIR {ftir_no}: Clicked #searchbtn")
    except Exception as e:
        logger.error(f"FTIR {ftir_no}: Could not click #searchbtn: {e}")
        return False

    # ── Step 7: Switch to FTIR detail popup window ─────────────────────
    detail_window = None
    start_wait = time.time()
    while time.time() - start_wait < 30:
        current_handles = driver.window_handles
        new_handles = [h for h in current_handles if h not in pre_search_windows]
        if new_handles:
            detail_window = new_handles[0]
            break
        # Also check window titles / URLs for FTIR detail indicators
        for h in current_handles:
            if h not in ([driver.window_handles[0]] if driver.window_handles else []) + [quick_search_window]:
                try:
                    driver.switch_to.window(h)
                    cur_url = driver.current_url.lower()
                    cur_title = driver.title.lower()
                    if "syqaa090" in cur_url or "ftir" in cur_title or len(driver.find_elements(By.ID, "PYQAA030E00S")) > 0:
                        detail_window = h
                        break
                except Exception:
                    continue
        if detail_window:
            break
        time.sleep(1)

    if not detail_window:
        logger.error(f"FTIR {ftir_no}: FTIR detail popup window did not open (invalid FTIR number or search timeout)")
        return False

    driver.switch_to.window(detail_window)
    logger.info(f"FTIR {ftir_no}: ✓ Switched to FTIR detail popup window ({detail_window})")
    time.sleep(2)
    return True


def search_and_extract_ftir(driver: webdriver.Edge, ftir_no: str, portal_url: Optional[str] = None) -> Dict[str, Any]:
    """
    Navigate to the FTIR detail page using the Quick Search popup chain,
    then extract attachment URLs from the resulting page.
    """
    if not portal_url:
        portal_url = PORTAL_SEARCH_URL or DEFAULT_SIFT_PORTAL_URL

    logger.info(f"Quick Search Triggered: ftir={ftir_no}")
    success = _navigate_to_ftir_detail_via_quick_search(driver, ftir_no, portal_url)

    if not success:
        logger.warning(f"FTIR {ftir_no}: Quick Search popup chain failed.")
        _close_extra_windows(driver)
        return {"page_url": driver.current_url, "subject_text": None, "attachment_urls": [], "page_title": ""}

    try:
        result = extract_ftir_page(driver, driver.current_url, ftir_no=ftir_no)
        return result
    except Exception as e:
        logger.error(f"FTIR {ftir_no}: Extraction from detail page failed: {e}")
        return {"page_url": driver.current_url, "subject_text": None, "attachment_urls": [], "page_title": ""}



# ---------------------------------------------------------------------------
#  FTIR Response Form Excel Download & Image Extraction
# ---------------------------------------------------------------------------

def _get_edge_download_dir(driver: webdriver.Edge) -> str:
    """Get or create a known download directory for Edge."""
    _project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    download_dir = os.path.join(_project_root, "temp_downloads")
    os.makedirs(download_dir, exist_ok=True)
    return download_dir


def _wait_for_download(download_dirs: Union[str, List[str]], timeout: int = 120) -> Optional[str]:
    """
    Wait for a new completed Excel file (.xlsx, .xls, .xlsm) to appear in download_dirs.
    Checks both the app's temp_downloads folder and the user's default ~/Downloads folder.
    """
    if isinstance(download_dirs, str):
        dirs = [download_dirs]
    else:
        dirs = list(download_dirs)
    
    # Add user's system Downloads folder as safe fallback
    user_downloads = os.path.expanduser("~/Downloads")
    if os.path.isdir(user_downloads) and user_downloads not in dirs:
        dirs.append(user_downloads)

    # Record baseline state of all directories
    baseline_files = {}
    for d in dirs:
        if os.path.isdir(d):
            try:
                baseline_files[d] = set(os.listdir(d))
            except Exception:
                baseline_files[d] = set()

    start_time = time.time()
    while time.time() - start_time < timeout:
        for d in dirs:
            if not os.path.isdir(d):
                continue
            try:
                current_files = set(os.listdir(d))
            except Exception:
                continue

            new_files = current_files - baseline_files.get(d, set())
            
            # Filter for completed Excel files
            completed_excel = [
                f for f in new_files
                if f.lower().endswith(('.xlsx', '.xls', '.xlsm'))
                and not f.endswith(('.crdownload', '.part', '.tmp', '.download'))
            ]
            
            if completed_excel:
                # Pick the newest file
                newest = max(
                    completed_excel,
                    key=lambda f: os.path.getmtime(os.path.join(d, f))
                )
                full_path = os.path.join(d, newest)
                
                # Verify file writing has stabilized
                prev_size = -1
                for _ in range(4):
                    time.sleep(0.5)
                    try:
                        curr_size = os.path.getsize(full_path)
                        if curr_size == prev_size and curr_size > 0:
                            logger.info(f"Download complete: {newest} ({curr_size:,} bytes) in {d}")
                            return full_path
                        prev_size = curr_size
                    except Exception:
                        pass
                return full_path

        time.sleep(1)

    logger.warning(f"Download wait timed out after {timeout}s across directories: {dirs}")
    return None


def _extract_images_from_xlsx(xlsx_path: str, output_dir: str) -> List[str]:
    """
    Extract embedded images from an Excel (.xlsx, .xlsm, or .xls) file.
    
    Uses a 4-tier fallback extraction strategy:
      1. OpenXML Zip extraction (inspects xl/media/ in standard .xlsx archives).
      2. openpyxl worksheet image inspection (ws._images).
      3. HTML / MHTML table image extraction (for portals that export HTML disguised as .xls).
      4. Binary stream carving (recovers embedded PNG/JPEG/BMP headers from legacy formats).
    """
    import zipfile
    import base64
    import re
    
    extracted = []
    os.makedirs(output_dir, exist_ok=True)
    
    # ── Strategy 1: OpenXML Zip extraction (Standard .xlsx) ────────────
    try:
        if zipfile.is_zipfile(xlsx_path):
            with zipfile.ZipFile(xlsx_path, 'r') as zf:
                media_files = [
                    f for f in zf.namelist()
                    if f.startswith('xl/media/') and not f.endswith('/')
                ]
                
                if media_files:
                    logger.info(f"Found {len(media_files)} embedded images in Excel OpenXML media")
                    for media_file in media_files:
                        filename = os.path.basename(media_file)
                        save_path = os.path.join(output_dir, filename)
                        
                        base, ext = os.path.splitext(save_path)
                        counter = 1
                        while os.path.exists(save_path):
                            save_path = f"{base}_{counter}{ext}"
                            counter += 1
                        
                        with zf.open(media_file) as src, open(save_path, 'wb') as dst:
                            dst.write(src.read())
                        
                        file_size = os.path.getsize(save_path)
                        logger.info(f"  Extracted image from Excel: {os.path.basename(save_path)} ({file_size:,} bytes)")
                        extracted.append(save_path)
                    
                    if extracted:
                        return extracted
    except Exception as e:
        logger.debug(f"Zip extraction attempt finished with: {e}")

    # ── Strategy 2: openpyxl worksheet image inspection ───────────────
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        images = getattr(ws, "_images", [])
        if images:
            logger.info(f"Found {len(images)} openpyxl image objects in worksheet")
            for idx, img in enumerate(images):
                img_ext = getattr(img, "format", "png") or "png"
                if not img_ext.startswith("."):
                    img_ext = f".{img_ext}"
                save_path = os.path.join(output_dir, f"openpyxl_img_{idx + 1}{img_ext}")
                
                # Get raw bytes from openpyxl image
                img_data = None
                if hasattr(img, "_data") and callable(img._data):
                    img_data = img._data()
                elif hasattr(img, "ref"):
                    img_data = img.ref.read() if hasattr(img.ref, "read") else None
                
                if img_data:
                    with open(save_path, "wb") as f:
                        f.write(img_data)
                    logger.info(f"  Extracted openpyxl image: {os.path.basename(save_path)}")
                    extracted.append(save_path)
            if extracted:
                return extracted
    except Exception as e:
        logger.debug(f"openpyxl image extraction attempt finished with: {e}")

    # ── Strategy 3: HTML / XML Spreadsheet data URL parsing ───────────
    try:
        with open(xlsx_path, "rb") as f:
            raw_bytes = f.read()
        
        # Check if file is HTML or XML disguised as Excel
        if b"<html" in raw_bytes[:1000].lower() or b"<?xml" in raw_bytes[:1000].lower() or b"<table" in raw_bytes[:1000].lower():
            logger.info("Excel file detected as HTML/XML format — scanning for inline images...")
            raw_str = raw_bytes.decode("utf-8", errors="ignore")
            
            # Find base64 image data URIs
            data_uri_pattern = r'data:image/(png|jpeg|jpg|gif|webp|bmp);base64,([A-Za-z0-9+/=]+)'
            matches = re.findall(data_uri_pattern, raw_str)
            for idx, (img_type, b64_str) in enumerate(matches):
                try:
                    img_bytes = base64.b64decode(b64_str)
                    save_path = os.path.join(output_dir, f"html_img_{idx + 1}.{img_type}")
                    with open(save_path, "wb") as f:
                        f.write(img_bytes)
                    logger.info(f"  Extracted inline base64 image: {os.path.basename(save_path)}")
                    extracted.append(save_path)
                except Exception:
                    continue
            if extracted:
                return extracted
    except Exception as e:
        logger.debug(f"HTML image extraction attempt finished with: {e}")

    # ── Strategy 4: Binary Stream Carving (PNG / JPEG / BMP) ───────────
    try:
        with open(xlsx_path, "rb") as f:
            content = f.read()
        
        # Look for PNG signature (\x89PNG\r\n\x1a\n) and IEND
        png_sig = b"\x89PNG\r\n\x1a\n"
        png_end = b"IEND\xaeB`\x82"
        pos = 0
        png_idx = 1
        while True:
            start = content.find(png_sig, pos)
            if start == -1:
                break
            end = content.find(png_end, start)
            if end != -1:
                end += len(png_end)
                png_bytes = content[start:end]
                if len(png_bytes) > 500:  # Ignore tiny icon fragments
                    save_path = os.path.join(output_dir, f"carved_img_{png_idx}.png")
                    with open(save_path, "wb") as f:
                        f.write(png_bytes)
                    extracted.append(save_path)
                    png_idx += 1
                pos = end
            else:
                pos = start + len(png_sig)

        # Look for JPEG signature (\xFF\xD8\xFF) and EOI (\xFF\xD9)
        jpg_sig = b"\xff\xd8\xff"
        jpg_end = b"\xff\xd9"
        pos = 0
        jpg_idx = 1
        while True:
            start = content.find(jpg_sig, pos)
            if start == -1:
                break
            end = content.find(jpg_end, start + 3)
            if end != -1:
                end += 2
                jpg_bytes = content[start:end]
                if len(jpg_bytes) > 1000:
                    save_path = os.path.join(output_dir, f"carved_img_{jpg_idx}.jpg")
                    with open(save_path, "wb") as f:
                        f.write(jpg_bytes)
                    extracted.append(save_path)
                    jpg_idx += 1
                pos = end
            else:
                pos = start + 3
                
        if extracted:
            logger.info(f"Carved {len(extracted)} image streams from Excel binary")
    except Exception as e:
        logger.debug(f"Binary carving finished with: {e}")
    
    return extracted
def _is_on_ftir_detail_page(driver: webdriver.Edge, switch_if_found: bool = True) -> bool:
    """
    Check if ANY open browser window is an FTIR detail page.
    If switch_if_found is True and a detail window is found but is not
    the current window, automatically switch to it.
    """
    def _check_current_window() -> bool:
        try:
            cur_url = driver.current_url.lower()
            cur_title = driver.title.lower()
            # Negative checks — these are NOT detail pages
            if "syqaa710" in cur_url or "menu" in cur_title:
                return False
            if "login" in cur_url or "auth" in cur_url:
                return False
            # A file download URL is not a detail page either
            if "ftirwebfile.do" in cur_url or "ftirwebthumbnail.do" in cur_url:
                return False
            # Positive checks
            if "syqaa090" in cur_url:
                return True
            if "ftir" in cur_title and "menu" not in cur_title:
                return True
            if len(driver.find_elements(By.ID, "PYQAA030E00S")) > 0:
                return True
        except Exception:
            pass
        return False

    # First check current window
    if _check_current_window():
        return True

    # If not found, scan ALL other windows
    if switch_if_found:
        try:
            original_handle = driver.current_window_handle
            for handle in driver.window_handles:
                if handle == original_handle:
                    continue
                try:
                    driver.switch_to.window(handle)
                    if _check_current_window():
                        logger.info(f"Found FTIR detail page in window: {handle}")
                        return True
                except Exception:
                    continue
            # If no detail window found, switch back to original
            driver.switch_to.window(original_handle)
        except Exception:
            pass

    return False



def extract_via_response_form(
    driver: webdriver.Edge,
    ftir_no: str,
    save_dir: str,
    url: Optional[str] = None,
) -> List[str]:
    """
    Strategy: Navigate to the FTIR Detail page, click the 'Go to FTIR Response Form'
    button (#PYQAA030E00S), which downloads the official Response Form Excel,
    and extract all embedded photos and metadata from it.
    """
    logger.info(f"FTIR {ftir_no}: [STRATEGY: RESPONSE FORM] Initiating Response Form download...")
    
    download_dir = _get_edge_download_dir(driver)
    
    # ── Step 1: Ensure browser is positioned on the FTIR detail window ──
    on_detail = _is_on_ftir_detail_page(driver)
    
    if not on_detail:
        # If url provided and contains detail page identifier, try direct link
        if url and url.startswith("http") and "syqaa090" in url.lower():
            try:
                logger.info(f"FTIR {ftir_no}: Trying direct hyperlink: {url}")
                driver.get(url)
                _wait_for_login_if_needed(driver)
                time.sleep(2)
                on_detail = _is_on_ftir_detail_page(driver)
            except Exception as e:
                logger.debug(f"Direct hyperlink load failed: {e}")
        
        # If still not on detail page (e.g. redirected to menu or direct URL failed)
        if not on_detail:
            logger.info(f"FTIR {ftir_no}: Navigating to FTIR detail window via Quick Search popup chain...")
            success = _navigate_to_ftir_detail_via_quick_search(driver, ftir_no)
            if not success:
                logger.warning(f"FTIR {ftir_no}: Could not navigate to FTIR detail window via Quick Search")
                _close_extra_windows(driver)
                return []
            on_detail = True

    # Configure Edge to download to our known directory (via CDP command on active window)
    try:
        driver.execute_cdp_cmd("Page.setDownloadBehavior", {
            "behavior": "allow",
            "downloadPath": os.path.abspath(download_dir),
        })
        logger.info(f"Download directory set via CDP to: {download_dir}")
    except Exception as e:
        logger.debug(f"CDP downloadPath configuration: {e}")

    time.sleep(2)

    button_strategies = [
        "//*[@id='PYQAA030E00S']",
        "//button[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'ftir response form')]",
        "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'ftir response form')]",
        "//input[contains(translate(@value, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'ftir response form')]",
        "//button[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'response form')]",
        "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'response form')]",
        "//input[contains(translate(@value, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'response form')]",
        "//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'go to ftir')]",
        "//*[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'response form')]",
    ]

    def _find_button_in_context():
        try:
            btn = driver.find_element(By.ID, "PYQAA030E00S")
            if btn.is_displayed() or btn.is_enabled():
                logger.info(f"✓ Found 'FTIR Response Form' button by exact ID #PYQAA030E00S")
                return btn
        except Exception:
            pass

        for xpath in button_strategies:
            try:
                elements = driver.find_elements(By.XPATH, xpath)
                for el in elements:
                    if el.is_displayed():
                        logger.info(f"Found button with text/value: '{el.text.strip() or el.get_attribute('value')}'")
                        return el
            except Exception:
                continue

        js_script = """
        function findButton(root) {
            var byId = document.getElementById('PYQAA030E00S');
            if (byId) return byId;
            let elements = Array.from(root.querySelectorAll('*'));
            for (let el of elements) {
                if (el.shadowRoot) {
                    let shadowBtn = findButton(el.shadowRoot);
                    if (shadowBtn) return shadowBtn;
                }
                let text = ((el.innerText || el.value || el.title || el.alt || el.name || '')).toLowerCase();
                if (text.includes('response form') || text.includes('go to ftir') || text.includes('ftir response')) {
                    if (el.offsetWidth > 0 || el.offsetHeight > 0) {
                        return el;
                    }
                }
            }
            return null;
        }
        return findButton(document);
        """
        try:
            btn = driver.execute_script(js_script)
            if btn:
                logger.info(f"JS brute-force search found button: {btn.tag_name}")
                return btn
        except Exception:
            pass
        return None

    def _recursive_iframe_search_and_click() -> bool:
        btn = _find_button_in_context()
        if btn:
            try:
                driver.execute_script("arguments[0].click();", btn)
                logger.info(f"FTIR {ftir_no}: Clicked 'FTIR Response Form' button using JavaScript. Waiting for download...")
                return True
            except Exception as e:
                logger.warning(f"JS click failed: {e}. Trying native click...")
                try:
                    btn.click()
                    logger.info(f"FTIR {ftir_no}: Clicked 'FTIR Response Form' button using native click. Waiting for download...")
                    return True
                except Exception as e2:
                    logger.warning(f"Native click also failed: {e2}")

        try:
            frames_tags = driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame")
        except Exception:
            return False

        for i in range(len(frames_tags)):
            try:
                current_frames = driver.find_elements(By.TAG_NAME, "iframe") + driver.find_elements(By.TAG_NAME, "frame")
                if i >= len(current_frames):
                    continue
                driver.switch_to.frame(current_frames[i])
                if _recursive_iframe_search_and_click():
                    return True
                driver.switch_to.parent_frame()
            except Exception:
                try:
                    driver.switch_to.default_content()
                except Exception:
                    pass
        return False

    driver.switch_to.default_content()
    clicked = _recursive_iframe_search_and_click()
    driver.switch_to.default_content()

    if not clicked:
        logger.warning(f"FTIR {ftir_no}: Could not find or click 'FTIR Response Form' button.")
        try:
            screenshot_path = os.path.join(os.getcwd(), f"debug_missing_button_{ftir_no}.png")
            driver.save_screenshot(screenshot_path)
        except Exception:
            pass
        _close_extra_windows(driver)
        return []

    # Wait for the Excel file to download (checks both temp_downloads and ~/Downloads)
    downloaded_file = _wait_for_download([download_dir, os.path.expanduser("~/Downloads")], timeout=120)

    if not downloaded_file:
        logger.warning(f"FTIR {ftir_no}: Response Form Excel download failed or timed out.")
        _close_extra_windows(driver)
        return []

    logger.info(f"FTIR {ftir_no}: Response Form Excel downloaded: {downloaded_file}")

    # Extract images from the Excel file
    extracted_images = _extract_images_from_xlsx(downloaded_file, save_dir)

    # Save a copy of the downloaded Excel into ftir_records/<ftir_no>/
    try:
        import shutil
        excel_dest = os.path.join(save_dir, f"{ftir_no}_response_form.xlsx")
        os.makedirs(save_dir, exist_ok=True)
        shutil.copy2(downloaded_file, excel_dest)
        logger.info(f"Saved Response Form Excel copy to: {excel_dest}")
    except Exception as e:
        logger.debug(f"Excel copy issue: {e}")

    # Clean up popup windows — return to main window
    _close_extra_windows(driver)

    if extracted_images:
        logger.info(f"FTIR {ftir_no}: ✓ Successfully extracted {len(extracted_images)} images from Response Form Excel!")
    else:
        logger.warning(f"FTIR {ftir_no}: Response Form Excel contained no extractable images.")

    return extracted_images

# ---------------------------------------------------------------------------
#  FTIR Response Form Excel Metadata Extraction
# ---------------------------------------------------------------------------

def extract_metadata_from_xlsx(xlsx_path: str) -> Dict[str, Any]:
    """
    Extract structured metadata from a downloaded FTIR Response Form Excel.

    Uses the exact cell-to-field map discovered from the proven reference
    script that works on the live SIFT portal.

    Parameters
    ----------
    xlsx_path : str
        Path to the downloaded Response Form Excel (.xlsx).

    Returns
    -------
    dict
        Dictionary of extracted fields, e.g.::
            {
                "ftir_number": "...",
                "subject": "...",
                "customer_complaint": "...",
                "diagnosis_code": "...",
                ...
            }
    """
    from openpyxl import load_workbook

    metadata = {}

    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active

        def _read_cells(row: int, cols: list, join_char: str = "") -> str:
            """Read multiple cells and join non-None values."""
            parts = [ws[f"{col}{row}"].value for col in cols]
            return join_char.join([str(p) for p in parts if p]).strip()

        wide_cols = ["D", "E", "F", "G", "H", "I", "J", "K"]
        narrow_cols = ["D", "E", "F"]
        date_cols = ["I", "J", "K"]

        metadata["ftir_number"]         = str(ws["C3"].value or "")
        metadata["subject"]             = _read_cells(8,  wide_cols, " ")
        metadata["product_model_code"]  = _read_cells(11, narrow_cols)
        metadata["sales_model_code"]    = _read_cells(12, narrow_cols)
        metadata["date_registered"]     = _read_cells(13, date_cols)
        metadata["date_of_incident"]    = _read_cells(14, date_cols)
        metadata["date_reported"]       = _read_cells(15, date_cols)
        metadata["casual_parts_number"] = _read_cells(15, narrow_cols)
        metadata["mileage"]             = _read_cells(16, date_cols)
        metadata["casual_parts_name"]   = _read_cells(16, narrow_cols, " ")
        metadata["days_used"]           = _read_cells(17, date_cols)
        metadata["customer_complaint"]  = _read_cells(21, wide_cols, " ")
        metadata["reproducibility"]     = _read_cells(22, wide_cols)
        metadata["incident_condition"]  = _read_cells(23, wide_cols, " ")
        metadata["checked_contents"]    = _read_cells(24, wide_cols, " ")
        metadata["checked_results"]     = _read_cells(25, wide_cols, " ")
        metadata["diagnosis_code"]      = _read_cells(26, wide_cols)
        metadata["repair_status"]       = _read_cells(29, wide_cols)
        metadata["repair_contents"]     = _read_cells(30, wide_cols, " ")
        metadata["problem_solved"]      = _read_cells(31, wide_cols)

        # Attachment file names listed in cells C43, C45, C47, C49, C51
        attachment_names = [ws[f"C{r}"].value for r in [43, 45, 47, 49, 51]]
        metadata["attachment_names"] = ",".join([str(a) for a in attachment_names if a])

        wb.close()
        logger.info(f"Extracted metadata from Response Form: FTIR={metadata.get('ftir_number')}, "
                     f"Subject={metadata.get('subject', '')[:50]}...")

    except Exception as e:
        logger.warning(f"Could not extract metadata from {xlsx_path}: {e}")

    return metadata


# ---------------------------------------------------------------------------
#  Attachment download
# ---------------------------------------------------------------------------

def _filename_from_response(response: requests.Response, url: str) -> str:
    """
    Derive a sensible filename from the HTTP response headers or URL.

    Priority:
    1. Content-Disposition header (``filename=...``)
    2. Last segment of the URL path
    3. Fallback to ``attachment`` + extension from Content-Type
    """
    # 1. Content-Disposition
    cd = response.headers.get("Content-Disposition", "")
    if "filename" in cd:
        # Handle both filename="name.ext" and filename*=UTF-8''name.ext
        match = re.search(r"filename\*?=['\"]?(?:UTF-8'')?([^'\";]+)", cd, re.IGNORECASE)
        if match:
            name = unquote(match.group(1)).strip()
            if name:
                return name

    # 2. URL path
    parsed = urlparse(url)
    path_segment = os.path.basename(parsed.path)
    if path_segment and "." in path_segment:
        return unquote(path_segment)

    # 3. Fallback: content-type → extension
    ct = response.headers.get("Content-Type", "application/octet-stream")
    ct = ct.split(";")[0].strip()
    ext = mimetypes.guess_extension(ct) or ".bin"
    return f"attachment{ext}"


def download_attachment(
    cookies: dict,
    url: str,
    save_dir: str,
    filename: Optional[str] = None,
    fallback_url: Optional[str] = None,
    timeout: int = 60,
    driver = None,
) -> Optional[str]:
    """
    Download a single attachment using the browser session's cookies,
    with automatic fallback to thumbnail URLs and in-browser fetch.

    Parameters
    ----------
    cookies : dict
        Cookie dict extracted from the Selenium session via
        ``_get_browser_cookies_for_requests(driver)``.
    url : str
        Direct URL of the attachment to download.
    save_dir : str
        Local directory to save the file into.
    filename : str, optional
        Explicit filename (e.g. baleno11.jpg). If None, derived from response/URL.
    fallback_url : str, optional
        Fallback URL (e.g. thumbnail endpoint) if primary URL fails.
    timeout : int
        HTTP request timeout in seconds.
    driver : webdriver.Edge, optional
        Active WebDriver instance for in-browser fetch fallback.

    Returns
    -------
    str or None
        Absolute path of the saved file, or None if download failed.
    """
    os.makedirs(save_dir, exist_ok=True)
    raw_content: Optional[bytes] = None
    target_filename = filename

    # ── Attempt 1 (BEST): In-Session Browser Fetch via JavaScript ──────
    # This preserves all SSO/Kerberos auth, session cookies, and CSRF tokens
    # that external requests.get() cannot replicate behind enterprise SSO.
    if driver is not None:
        logger.info(f"Attempting in-session browser fetch for: {url}")
        raw_content = _download_via_browser_fetch(driver, url)
        if not raw_content and fallback_url:
            logger.info(f"Attempting in-session browser fetch for fallback: {fallback_url}")
            raw_content = _download_via_browser_fetch(driver, fallback_url)

    # ── Attempt 2: Direct HTTP GET on Primary URL ──────────────────────
    if not raw_content:
        try:
            resp = requests.get(url, cookies=cookies, timeout=timeout)
            ct = resp.headers.get("Content-Type", "").lower()
            if resp.status_code == 200 and len(resp.content) > 100 and "text/html" not in ct:
                raw_content = resp.content
                if not target_filename:
                    target_filename = _filename_from_response(resp, url)
        except Exception as e:
            logger.debug(f"Direct request failed for {url}: {e}")

    # ── Attempt 3: Fallback URL HTTP GET (e.g. Thumbnail) ──────────────
    if not raw_content and fallback_url:
        try:
            resp_fb = requests.get(fallback_url, cookies=cookies, timeout=timeout)
            ct_fb = resp_fb.headers.get("Content-Type", "").lower()
            if resp_fb.status_code == 200 and len(resp_fb.content) > 100 and "text/html" not in ct_fb:
                raw_content = resp_fb.content
                if not target_filename:
                    target_filename = _filename_from_response(resp_fb, fallback_url)
        except Exception as e:
            logger.debug(f"Fallback request failed for {fallback_url}: {e}")

    # ── Attempt 4: Selenium Element Screenshot Fallback (if HTML viewer) ─
    if not raw_content and driver is not None:
        try:
            original_window = driver.current_window_handle
            driver.switch_to.new_window('tab')
            driver.get(url)
            time.sleep(2)
            imgs = driver.find_elements(By.TAG_NAME, "img")
            largest_img = None
            max_area = 0
            for img in imgs:
                try:
                    size = img.size
                    area = size["width"] * size["height"]
                    if area > max_area:
                        max_area = area
                        largest_img = img
                except Exception:
                    pass
            if largest_img and max_area > 5000:
                if not target_filename:
                    target_filename = "viewer_screenshot.png"
                save_path = os.path.join(save_dir, target_filename)
                base, ext = os.path.splitext(save_path)
                counter = 1
                while os.path.exists(save_path):
                    save_path = f"{base}_{counter}{ext}"
                    counter += 1
                largest_img.screenshot(save_path)
                logger.info(f"  Screenshot extracted from viewer: {os.path.basename(save_path)}")
                return save_path
        except Exception as e:
            logger.debug(f"Viewer screenshot fallback issue: {e}")
        finally:
            try:
                driver.close()
                driver.switch_to.window(original_window)
            except Exception:
                pass

    if not raw_content or len(raw_content) < 100:
        logger.warning(f"Failed to download valid image from {url}")
        return None

    # Derive sensible filename
    if not target_filename:
        parsed = urlparse(url)
        path_base = os.path.basename(parsed.path)
        if "." in path_base:
            target_filename = path_base
        else:
            target_filename = "attachment.jpg"

    # Sanitize filename
    target_filename = re.sub(r'[<>:"/\\|?*]', "_", target_filename)
    if not any(target_filename.lower().endswith(ext) for ext in [".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp", ".pdf", ".xlsx"]):
        target_filename += ".jpg"

    save_path = os.path.join(save_dir, target_filename)
    base, ext = os.path.splitext(save_path)
    counter = 1
    while os.path.exists(save_path):
        save_path = f"{base}_{counter}{ext}"
        counter += 1

    with open(save_path, "wb") as f:
        f.write(raw_content)

    file_size = os.path.getsize(save_path)
    logger.info(f"  ✓ Saved attachment: {os.path.basename(save_path)} ({file_size:,} bytes)")
    return save_path


# ---------------------------------------------------------------------------
#  End-to-end single-FTIR processor
# ---------------------------------------------------------------------------

def process_ftir(
    driver: webdriver.Edge,
    ftir_no: str,
    url: str,
    base_save_dir: str = "ftir_records",
) -> Dict[str, Any]:
    """
    Extract and download all attachments for a single FTIR record.

    Creates ``<base_save_dir>/<ftir_no>/`` and downloads every attachment
    found on the FTIR detail page into it.
    """
    save_dir = os.path.join(base_save_dir, str(ftir_no))

    # ── Force Fresh Download ───────────────────────────────────────────
    if os.path.isdir(save_dir):
        logger.info(f"FTIR {ftir_no}: Clearing existing cached files for fresh download")
        for f in os.listdir(save_dir):
            file_path = os.path.join(save_dir, f)
            if os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                except OSError as e:
                    logger.warning(f"Could not remove cached file {file_path}: {e}")

    extraction_source = "None"
    attachment_urls = []
    downloaded: List[str] = []
    page_info = {"subject_text": None, "ftir_metadata": {}}

    # ── Ensure we are on the FTIR Detail Page ──────────────────────────
    # _is_on_ftir_detail_page scans ALL windows and auto-switches if found
    if not _is_on_ftir_detail_page(driver, switch_if_found=True):
        logger.info(f"FTIR {ftir_no}: Navigating to detail page via Quick Search...")
        _navigate_to_ftir_detail_via_quick_search(driver, ftir_no)
    else:
        logger.info(f"FTIR {ftir_no}: ✓ Already on FTIR detail page")

    # ── PRIMARY STRATEGY: FTIR Response Form Excel Download ──────────
    logger.info(f"FTIR {ftir_no}: [PRIMARY STRATEGY] Attempting FTIR Response Form Excel extraction...")
    try:
        extracted_images = extract_via_response_form(driver, ftir_no, save_dir, url=url)
        if extracted_images:
            extraction_source = "Excel Response Form"
            downloaded = extracted_images
            logger.info(f"FTIR {ftir_no}: ✓ Response Form strategy succeeded with {len(downloaded)} image(s)")
    except Exception as e:
        logger.warning(f"FTIR {ftir_no}: Excel Response Form extraction exception: {e}")

    # Always extract page info for metadata (and secondary fallback)
    try:
        page_info = extract_ftir_page(driver, url or driver.current_url, ftir_no=ftir_no)
        attachment_items = page_info.get("attachment_items", [])
        attachment_urls = page_info.get("attachment_urls", [])
    except Exception as e:
        logger.warning(f"FTIR {ftir_no}: Live DOM metadata extraction exception: {e}")
        attachment_items = []
        attachment_urls = []

    # ── SECONDARY STRATEGY: Live SIFT Detail Page DOM & Entity Extraction ─
    if not downloaded and attachment_items:
        logger.info(f"FTIR {ftir_no}: [SECONDARY STRATEGY] Extracting live DOM entities & photos...")
        cookies = _get_browser_cookies_for_requests(driver)
        for i, item in enumerate(attachment_items):
            logger.info(f"FTIR {ftir_no}: downloading attachment {i + 1}/{len(attachment_items)}")
            try:
                saved = download_attachment(
                    cookies=cookies,
                    url=item["url"],
                    save_dir=save_dir,
                    filename=item.get("filename"),
                    fallback_url=item.get("fallback_url"),
                    driver=driver,
                )
                if saved:
                    downloaded.append(saved)
            except Exception as e:
                logger.error(f"Error downloading attachment {item.get('url')}: {e}")
            if i < len(attachment_items) - 1:
                time.sleep(_REQUEST_DELAY)

        if downloaded:
            extraction_source = "SIFT Live Detail Page"
            logger.info(f"FTIR {ftir_no}: ✓ Live detail page strategy succeeded with {len(downloaded)} image(s)")

    # ── Extract structured metadata from saved Response Form Excel if available ─
    ftir_metadata = page_info.get("ftir_metadata", {})
    response_form_path = os.path.join(save_dir, f"{ftir_no}_response_form.xlsx")
    if os.path.isfile(response_form_path):
        try:
            excel_meta = extract_metadata_from_xlsx(response_form_path)
            for k, v in excel_meta.items():
                if v and not ftir_metadata.get(k):
                    ftir_metadata[k] = v
            if not page_info["subject_text"] and ftir_metadata.get("subject"):
                page_info["subject_text"] = ftir_metadata["subject"]
        except Exception as e:
            logger.warning(f"FTIR {ftir_no}: Metadata extraction from Excel failed: {e}")

    # ── Clean up leftover popup windows ────────────────────────────────
    _close_extra_windows(driver)

    logger.info(
        f"FTIR {ftir_no}: {len(downloaded)} attachments saved to {save_dir} (Source: {extraction_source})"
    )

    return {
        "ftir_no": ftir_no,
        "url": url,
        "save_dir": save_dir,
        "subject_text": page_info.get("subject_text"),
        "downloaded_files": downloaded,
        "attachment_urls": attachment_urls,
        "skipped": False,
        "extraction_source": extraction_source,
        "ftir_metadata": ftir_metadata,
    }


# ---------------------------------------------------------------------------
#  CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )

    parser = argparse.ArgumentParser(
        description="Test FTIR attachment extraction against a single page URL.",
    )
    parser.add_argument(
        "url",
        help="Full URL of an FTIR detail page to extract attachments from.",
    )
    parser.add_argument(
        "--ftir-no",
        default="TEST_FTIR",
        help="FTIR record number / folder name (default: TEST_FTIR).",
    )
    parser.add_argument(
        "--profile",
        default=None,
        help=f"Chrome user-data-dir profile path (default: {_DEFAULT_PROFILE_DIR}).",
    )
    parser.add_argument(
        "--save-dir",
        default="ftir_records",
        help="Root directory for downloaded attachments (default: ftir_records).",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run Chrome in headless mode (no visible window).",
    )

    args = parser.parse_args()

    print("=" * 60)
    print("  FTIR Browser Extraction — Single-Page Test")
    print("=" * 60)
    print(f"  URL        : {args.url}")
    print(f"  FTIR No    : {args.ftir_no}")
    print(f"  Profile    : {args.profile or _DEFAULT_PROFILE_DIR}")
    print(f"  Save Dir   : {args.save_dir}")
    print(f"  Headless   : {args.headless}")
    print("=" * 60)

    drv = get_driver(profile_dir=args.profile, headless=args.headless)

    try:
        result = process_ftir(
            driver=drv,
            ftir_no=args.ftir_no,
            url=args.url,
            base_save_dir=args.save_dir,
        )

        print("\n" + "=" * 60)
        print("  Results")
        print("=" * 60)
        print(f"  Skipped (cached)  : {result['skipped']}")
        print(f"  Subject text      : {result['subject_text'] or '(not found on page)'}")
        print(f"  Attachment URLs   : {len(result['attachment_urls'])}")
        for u in result["attachment_urls"]:
            print(f"    → {u}")
        print(f"  Downloaded files  : {len(result['downloaded_files'])}")
        for f in result["downloaded_files"]:
            print(f"    → {f}")
        print(f"  Save directory    : {result['save_dir']}")
        print("=" * 60)

    finally:
        drv.quit()
        print("\nBrowser closed.")
