"""
Excel Input/Output Module (src.excel_io)
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

Reads incoming FTIR Excel spreadsheets using openpyxl (for hyperlink extraction)
and pandas (for convenient DataFrame handling), then merges hyperlink targets
into the DataFrame as an ``ftir_url`` column.

Also provides write_output_sheet() for formatting the final automated
defect report at the end of the pipeline.

Usage::

    python -m src.excel_io data/excel/some_file.xlsx
"""

import sys
import os
import logging
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
#  Helpers
# ---------------------------------------------------------------------------

# Known header keywords that signal the actual header row in sheets
# where the first few rows contain metadata / titles.
_HEADER_MARKERS = {
    "masked model", "model", "vin", "masked vin", "fc-ok",
    "subject (english)", "subject", "date registered", "mileage",
    "sno", "ftir", "reported country",
}


def _detect_header_row(ws) -> int:
    """
    Walk down the first 10 rows of a worksheet looking for the real
    header row.  Returns the 1-based openpyxl row index that contains
    the most recognised column header keywords.  Falls back to row 1.
    """
    best_row = 1
    best_score = 0
    for row_idx in range(1, min(ws.max_row + 1, 11)):
        row_vals = [
            str(ws.cell(row=row_idx, column=c).value or "").strip().lower()
            for c in range(1, ws.max_column + 1)
        ]
        score = sum(1 for v in row_vals if v in _HEADER_MARKERS)
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row


# ---------------------------------------------------------------------------
#  Read
# ---------------------------------------------------------------------------

def read_ftir_sheet(path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Load an FTIR Excel spreadsheet and extract hyperlinks from cells.

    1. Opens the workbook with **openpyxl** to auto-detect the real header
       row (some files have metadata / title rows above the data).
    2. Walks every data cell collecting ``cell.hyperlink.target`` values.
    3. Reads the same range with **pandas** ``read_excel`` so every column
       comes through with proper dtype handling.
    4. Adds a synthetic ``ftir_url`` column.  For each row the value is the
       *first* hyperlink found in that row (scanning left-to-right), which
       in practice is the FTIR number column.  If no hyperlink is present
       the value is ``None``.

    Parameters
    ----------
    path : str
        Filesystem path to the input ``.xlsx`` file.
    sheet_name : str, optional
        Worksheet name to read.  Defaults to the active sheet.

    Returns
    -------
    pd.DataFrame
        All original columns preserved as-is, plus ``ftir_url``.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"FTIR spreadsheet not found: {path}")

    # ------------------------------------------------------------------
    # Step 1 — openpyxl: detect header row & extract hyperlinks
    # ------------------------------------------------------------------
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name] if sheet_name else wb.active
    resolved_sheet = ws.title

    header_row = _detect_header_row(ws)
    data_start = header_row + 1  # first data row in openpyxl (1-based)

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]

    # Build hyperlink map: data_row_0based -> first hyperlink target
    hyperlinks: dict[int, str] = {}

    for row_idx in range(data_start, ws.max_row + 1):
        pandas_idx = row_idx - data_start  # 0-based index matching pandas
        first_link = None
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.hyperlink and cell.hyperlink.target:
                first_link = cell.hyperlink.target
                break  # first hyperlink per row is enough
            elif isinstance(cell.value, str) and cell.value.strip().startswith("http"):
                first_link = cell.value.strip()
                break

        if first_link is not None:
            hyperlinks[pandas_idx] = first_link

    wb.close()

    # ------------------------------------------------------------------
    # Step 2 — pandas: read tabular data with proper dtypes
    # ------------------------------------------------------------------
    # header= argument is 0-based row index to use as header
    # skiprows skips everything above the header row
    skip = header_row - 1  # rows to skip before the header
    df = pd.read_excel(path, sheet_name=resolved_sheet, header=0, skiprows=skip if skip > 0 else None)

    # Drop completely empty rows (sometimes trailing blanks exist)
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    # ------------------------------------------------------------------
    # Step 3 — merge hyperlinks into the DataFrame
    # ------------------------------------------------------------------
    df["ftir_url"] = df.index.map(lambda i: hyperlinks.get(i))

    link_count = sum(1 for v in df["ftir_url"] if v is not None)
    logging.info(
        f"Loaded {len(df)} FTIR records from '{os.path.basename(path)}' "
        f"(header row {header_row}, {link_count} rows have hyperlinks)"
    )

    return df


# ---------------------------------------------------------------------------
#  Write
# ---------------------------------------------------------------------------

def write_output_sheet(df: pd.DataFrame, path: str) -> None:
    """
    Write a DataFrame to a professionally formatted Excel report.

    Generates an ``.xlsx`` file with:
    * Bold white-on-dark-blue header row.
    * Alternating row shading for readability.
    * Auto-adjusted column widths based on content length.
    * ``Flag_For_Review`` column cells highlighted in yellow when ``True``.

    Parameters
    ----------
    df : pd.DataFrame
        Processed evaluation dataframe containing model predictions and
        fusion decisions.
    path : str
        Destination filesystem path for the generated report (``.xlsx``).
    """
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FTIR Triage Report"

    # -- Styles --
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    flag_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    columns = list(df.columns)

    # 1. Hide ftir_url column from the final output, but keep it available for links
    has_ftir_url = "ftir_url" in columns
    if has_ftir_url:
        columns.remove("ftir_url")
        
    # Find which column holds the FTIR number (so we can attach the link to it)
    ftir_col_idx = None
    for idx, col in enumerate(columns):
        col_lower = str(col).lower()
        if "ftir" in col_lower and ("no" in col_lower or "num" in col_lower):
            ftir_col_idx = idx + 1
            break

    # -- Header row --
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # -- Data rows --
    flag_col_idx = None
    if "Flag_For_Review" in columns:
        flag_col_idx = columns.index("Flag_For_Review") + 1

    # Openpyxl hyperlink style (blue, underlined)
    link_font = Font(color="0563C1", underline="single")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = row[col_name]
            # Convert numpy/pandas types to native Python for openpyxl
            if pd.isna(value):
                value = None
            elif hasattr(value, "item"):
                value = value.item()

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Alternating row shading
            if row_idx % 2 == 0:
                cell.fill = alt_fill

            # If this is the FTIR number column, attach the hidden ftir_url!
            if col_idx == ftir_col_idx and has_ftir_url:
                url = row.get("ftir_url")
                if url and isinstance(url, str) and url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = link_font
                    # If shading was applied, preserve it, otherwise it's just link font


        # Highlight flagged rows
        if flag_col_idx:
            flag_cell = ws.cell(row=row_idx, column=flag_col_idx)
            if flag_cell.value is True or str(flag_cell.value).strip().lower() == "true":
                flag_cell.fill = flag_fill
                flag_cell.font = Font(bold=True, color="CC0000")

    # -- Auto-width columns --
    for col_idx in range(1, len(columns) + 1):
        max_len = len(str(columns[col_idx - 1]))  # header length
        for row_idx in range(2, min(ws.max_row + 1, 102)):  # sample first 100 rows
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        # Cap at 50 characters width, add padding
        adjusted_width = min(max_len + 3, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # -- Freeze header row --
    ws.freeze_panes = "A2"

    wb.save(path)
    wb.close()

    logging.info(f"Output report saved to {path} ({len(df)} records)")


# ---------------------------------------------------------------------------
#  CLI verification entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

    if len(sys.argv) < 2:
        # Default: try each xlsx found in data/excel/
        excel_dir = os.path.join("data", "excel")
        xlsx_files = sorted(f for f in os.listdir(excel_dir) if f.endswith(".xlsx")) if os.path.isdir(excel_dir) else []
        if not xlsx_files:
            print("Usage: python -m src.excel_io <path_to_ftir.xlsx>")
            sys.exit(1)
        sheet_path = os.path.join(excel_dir, xlsx_files[0])
        print(f"No path given — auto-detected: {sheet_path}\n")
    else:
        sheet_path = sys.argv[1]

    df = read_ftir_sheet(sheet_path)

    print(f"\nShape: {df.shape}")
    print(f"Columns: {list(df.columns)}\n")

    # Show first 5 rows with all columns
    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 200)
    pd.set_option("display.max_colwidth", 45)
    print(df.head(5).to_string())

    # Highlight the ftir_url column specifically
    print("\n--- ftir_url column (first 10 rows) ---")
    for i, url in enumerate(df["ftir_url"].head(10)):
        print(f"  Row {i}: {url}")

    # Quick round-trip test of write_output_sheet
    test_out = os.path.join("outputs", "_test_roundtrip.xlsx")
    write_output_sheet(df.head(5), test_out)
    print(f"\nRound-trip write test saved to: {test_out}")
